from pathlib import Path

import openpyxl


def create_template(original_file: Path, template_file: Path, header_rows: int = 2):
    wb = openpyxl.load_workbook(original_file)

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]

        max_col = sheet.max_column

        for row in range(sheet.max_row, header_rows, -1):
            for col in range(1, max_col + 1):
                sheet.cell(row, col).value = None

    wb.save(template_file)
    print(f"Шаблон сохранён: {template_file}")


def main() -> None:
    file = Path("../../private/64931.xlsx")

    create_template(
        file,
        Path("article_template.xlsx"),
        header_rows=2,
    )

    wb = openpyxl.load_workbook(file)

    sheets = wb.sheetnames
    for sheetname in sheets:
        sheet = wb[sheetname]

        merges = []
        for range in sheet.merged_cells.ranges:
            merges.append(range.bounds)

        print(merges)

        headers = {}
        for row, row_obj in enumerate(sheet.values):
            if row > 1:
                break
            for col, val in enumerate(row_obj):
                if val is None:
                    continue
                headers[(row, col)] = val
        print(headers)


if __name__ == "__main__":
    main()
