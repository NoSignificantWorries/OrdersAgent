from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import xlrd

from . import config as conf
from . import functional as func


@dataclass
class ProductPosition:
    material: str
    x: int
    y: int
    amount: int
    barcode: str | None = None
    marking: str | None = None


@dataclass
class ParseResults:
    unique_materials: dict[str, list[int]]
    data: list[ProductPosition]
    count: int


@dataclass
class CellValue:
    value: str | int | tuple[int, int]
    type: conf.CellType = conf.CellType.TEXT
    merged: bool = False
    parent: tuple[int, int] | None = None


@dataclass
class HeaderCell:
    type: conf.CellType
    row: int
    col: int
    merged: bool = False
    parent: tuple[int, int] | None = None

    def to_json(self) -> dict[str, str | int | tuple[int, int] | bool | None]:
        return {
            "type": self.type.value,
            "type-name": self.type.name,
            "row": self.row,
            "col": self.col,
            "merged": self.merged,
            "parent": self.parent
        }


def cell_value_classify(value: str) -> tuple[conf.CellType, list[str] | None]:
    if conf.TYPES_CONFIG.regex is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.regex.items():
            for pattern, groups in patterns:
                matched, values = func.get_match_and_groups(pattern, value, groups)
                if matched:
                    return cell_type, values
    if conf.TYPES_CONFIG.fuzzy is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.fuzzy.items():
            for pattern in patterns:
                if func.fuzzy_match(value.lower(), pattern, 70):
                    return cell_type, None
    return conf.CellType.TEXT, None


class Sheet:
    def __init__(self, nrows: int, ncols: int, name: str) -> None:
        self.nrows: int = nrows
        self.ncols: int = ncols
        self.name: str = name
        self.data: list[list[CellValue | None]] = [[None] * ncols for _ in range(nrows)]

        self.headers: list[HeaderCell] = []

        self._empty_rows: list[int] = [i for i in range(nrows)]
        self._empty_cols: list[int] = [i for i in range(ncols)]

    @property
    def empty(self) -> bool:
        return self.nrows == 0 or self.ncols == 0

    def _check_position(self, row: int, col: int) -> None:
        if row in self._empty_rows:
            idx = self._empty_rows.index(row)
            del self._empty_rows[idx]

        if col in self._empty_cols:
            idx = self._empty_cols.index(col)
            del self._empty_cols[idx]

    def add_cell(self, row: int, col: int, cell: CellValue | None) -> None:
        if cell is not None:
            self.data[row][col] = cell
            if not cell.merged:
                self._check_position(row, col)

    def get_cell(self, row: int, col: int) -> CellValue | None:
        return self.data[row][col]

    def clean_empty_cols_and_rows(self) -> None:
        for i in self._empty_rows[::-1]:
            del self.data[i]
        self._empty_rows = []

        for i in range(len(self.data)):
            for j in self._empty_cols[::-1]:
                del self.data[i][j]
        self._empty_cols = []

        self.nrows = len(self.data)
        if self.nrows > 0:
            self.ncols = len(self.data[0])
        else:
            self.ncols = 0

    def find_headers(self) -> list[HeaderCell]:
        for irow, row in enumerate(self.data):
            for icol, cell in enumerate(row):
                if cell is None:
                    continue
                if cell.type in conf.HEADER:
                    self.headers.append(HeaderCell(type=cell.type, row=irow, col=icol, merged=cell.merged, parent=cell.parent))
        return self.headers


class Document:
    def __init__(self, name: str | None = None) -> None:
        self.name: str | None = name
        self.sheets: dict[str, Sheet] = {}

    def add_sheet(self, nrows: int, ncols: int, name: str) -> None:
        sheet = Sheet(nrows=nrows, ncols=ncols, name=name)
        self.sheets[name] = sheet

    def add_cell_on_sheet(self, sheetname: str, row: int, col: int, cell: CellValue | None) -> None:
        sheet = self.sheets.get(sheetname)
        if sheet is None:
            raise ValueError(f"Sheet with name '{sheetname}' is not exists in the document '{self.name}'")

        sheet.add_cell(row=row, col=col, cell=cell)

    def get_cell_from_sheet(self, sheetname: str, row: int, col: int) -> CellValue | None:
        sheet = self.sheets.get(sheetname)
        if sheet is None:
            raise ValueError(f"Sheet with name '{sheetname}' is not exists in the document '{self.name}'")
        return sheet.get_cell(row, col)

    def clean_sheets(self) -> None:
        for sheet in self.sheets.values():
            sheet.clean_empty_cols_and_rows()

    def iter_sheets(self) -> Iterator[tuple[str, Sheet]]:
        yield from self.sheets.items()


class TableLoader:
    @staticmethod
    def load(bdata: BytesIO | None = None, filepath: Path | None = None) -> Document | None:
        '''
        Loads data from xls/xlsx files by the provided path or from bytes format.

        Args:
            bdata (None or BytesIO object): Bytes data of a file.
            filepath (PosixPath or None): Path to the target file.

        Returns:
            ---

        Raises:
            ValueError: If file type is unsupported or all arguments is None.
        '''

        if filepath:
            format = filepath.suffix.lower()
            if format == ".xls":
                return TableLoader._load_xls_data(filepath=filepath)
            elif format == ".xlsx":
                return TableLoader._load_xlsx_data(filepath=filepath)
            else:
                raise ValueError("Unsupported file type")
        elif bdata:
            bdata.seek(0)
            header = bdata.read(8)
            bdata.seek(0)

            if header.startswith(b"PK\x03\x04"):
                return TableLoader._load_xlsx_data(bdata=bdata)
            else:
                return TableLoader._load_xls_data(bdata=bdata)
        raise ValueError("No data provided")

    @staticmethod
    def _parse_cell(value: Any) -> None | CellValue:
        clean_text = func.to_text(value)
        if clean_text is None:
            return None

        number_value = func.number_to_int(clean_text)
        if number_value is not None:
            return CellValue(value=number_value, type=conf.CellType.NUMBER)

        cell_type, parsed_values = cell_value_classify(clean_text)

        if cell_type is conf.CellType.SIZES:
            x = func.number_to_int(parsed_values[0])
            y = func.number_to_int(parsed_values[1])
            return CellValue(value=(x, y), type=cell_type)

        return CellValue(value=str(value), type=cell_type)

    @staticmethod
    def _get_merged_cells_from_xls_sheet(sheet) -> dict[tuple[int, int], tuple[int, int]]:
        merged_cells: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in sheet.merged_cells:
            parent = (merged_range[0], merged_range[2])
            for row_idx in range(merged_range[0], merged_range[1]):
                for col_idx in range(merged_range[2], merged_range[3]):
                    merged_cells[(row_idx, col_idx)] = parent
            del merged_cells[parent]
        return merged_cells

    @staticmethod
    def _load_xls_data(
        bdata: BytesIO | None = None, filepath: Path | None = None
    ) -> Document | None:
        try:
            if filepath:
                wb = xlrd.open_workbook(str(filepath), formatting_info=True)
            elif bdata:
                bdata.seek(0)
                wb = xlrd.open_workbook(
                    file_contents=bdata.read(), formatting_info=True
                )
            else:
                raise ValueError("Unsupported file type")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = xlrd.open_workbook(str(filepath), formatting_info=False)
                elif bdata:
                    bdata.seek(0)
                    wb = xlrd.open_workbook(
                        file_contents=bdata.read(), formatting_info=False
                    )
                else:
                    raise ValueError("Unsupported file type")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return

        document = Document()
        sheets = wb.sheet_names()
        for sheetname in sheets:
            sheet = wb[sheetname]
            document.add_sheet(sheet.nrows, sheet.ncols, sheetname)

            merged_cells = {}
            if hasattr(sheet, "merged_cells"):
                merged_cells = TableLoader._get_merged_cells_from_xls_sheet(sheet)
            # print(merged_cells)

            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)
                    cell_data = TableLoader._parse_cell(value)

                    merge_parent = merged_cells.get((row, col))
                    if merge_parent and cell_data is None and merge_parent != (row, col):
                        parent_cell = document.get_cell_from_sheet(sheetname, *merge_parent)
                        cell_data = parent_cell
                        if cell_data is not None:
                            cell_data.parent = merge_parent
                            cell_data.merged = True

                    document.add_cell_on_sheet(sheetname, row, col, cell_data)
        document.clean_sheets()
        return document


    @staticmethod
    def _get_merged_cells_from_xlsx_sheet(sheet) -> dict[tuple[int, int], tuple[int, int]]:
        merged_cells: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in sheet.merged_cells.ranges:
            parent = (merged_range.min_row - 1, merged_range.min_col - 1)
            for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                    merged_cells[(row_idx, col_idx)] = parent
            del merged_cells[parent]
        return merged_cells

    @staticmethod
    def _load_xlsx_data(
        bdata: BytesIO | None = None, filepath: Path | None = None
    ) -> Document | None:
        try:
            if filepath:
                wb = openpyxl.load_workbook(filepath, data_only=False)
            elif bdata:
                bdata.seek(0)
                wb = openpyxl.load_workbook(bdata, data_only=False)
            else:
                raise ValueError("No data proveded")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                elif bdata:
                    bdata.seek(0)
                    wb = openpyxl.load_workbook(bdata, data_only=True)
                else:
                    raise ValueError("No data proveded")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return

        document = Document()
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]
            document.add_sheet(sheet.max_row - 1, sheet.max_column - 1, sheetname)

            merged_cells = {}
            if with_metadata:
                merged_cells = TableLoader._get_merged_cells_from_xlsx_sheet(sheet)

            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    cell_data = TableLoader._parse_cell(value)

                    merge_parent = merged_cells.get((cell.row - 1, cell.column - 1))
                    if merge_parent and cell_data is None and merge_parent != (cell.row - 1, cell.column - 1):
                        parent_cell = document.get_cell_from_sheet(sheetname, *merge_parent)
                        cell_data = parent_cell
                        if cell_data is not None:
                            cell_data.parent = merge_parent
                            cell_data.merged = True
                    document.add_cell_on_sheet(sheetname, cell.row - 1, cell.col - 1, cell_data)
        document.clean_sheets()
        return document
