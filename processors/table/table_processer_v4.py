from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import openpyxl
import xlrd

from .config import CellType


@dataclass
class ProductPosition:
    material: str
    x: int
    y: int
    amount: int
    barcode: str | None = None
    marking: str | None = None


@dataclass
class ParseResults:
    unique_materials: dict[str, list[int]]
    data: list[ProductPosition]
    count: int


class TableLoader:
    @staticmethod
    def load(bdata: BytesIO | None = None, filepath: Path | None = None):
        '''
        Loads data from xls/xlsx files by the provided path or from bytes format.

        Args:
            bdata (None or BytesIO object): Bytes data of a file.
            filepath (PosixPath or None): Path to the target file.

        Returns:
            ---

        Raises:
            ValueError: If file type is unsupported or all arguments is None.
        '''

        if filepath:
            format = filepath.suffix.lower()
            if format == ".xls":
                return TableLoader._load_xls_data(filepath=filepath)
            elif format == ".xlsx":
                return TableLoader._load_xlsx_data(filepath=filepath)
            else:
                raise ValueError("Unsupported file type")
        elif bdata:
            bdata.seek(0)
            header = bdata.read(8)
            bdata.seek(0)

            if header.startswith(b"PK\x03\x04"):
                return TableLoader._load_xlsx_data(bdata=bdata)
            else:
                return TableLoader._load_xls_data(bdata=bdata)
        raise ValueError("No data provided")


    @staticmethod
    def _get_merged_cells_from_xls_sheet(sheet) -> dict[tuple[int, int], tuple[int, int]]:
        merged_cells: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in sheet.merged_cells:
            parent = (merged_range[0], merged_range[2])
            for row_idx in range(merged_range[0], merged_range[1]):
                for col_idx in range(merged_range[2], merged_range[3]):
                    merged_cells[(row_idx, col_idx)] = parent
        return merged_cells

    @staticmethod
    def _load_xls_data(
        bdata: BytesIO | None = None, filepath: Path | None = None
    ):
        try:
            if filepath:
                wb = xlrd.open_workbook(str(filepath), formatting_info=True)
            elif bdata:
                bdata.seek(0)
                wb = xlrd.open_workbook(
                    file_contents=bdata.read(), formatting_info=True
                )
            else:
                raise ValueError("Unsupported file type")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = xlrd.open_workbook(str(filepath), formatting_info=False)
                elif bdata:
                    bdata.seek(0)
                    wb = xlrd.open_workbook(
                        file_contents=bdata.read(), formatting_info=False
                    )
                else:
                    raise ValueError("Unsupported file type")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheets = wb.sheet_names()
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_cells = {}
            if hasattr(sheet, "merged_cells"):
                merged_cells = TableLoader._get_merged_cells_from_xls_sheet(sheet)

            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)


    @staticmethod
    def _get_merged_cells_from_xlsx_sheet(sheet) -> dict[tuple[int, int], tuple[int, int]]:
        merged_cells: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in sheet.merged_cells.ranges:
            parent = (merged_range.min_row - 1, merged_range.min_col - 1)
            for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                    merged_cells[(row_idx, col_idx)] = parent
        return merged_cells

    @staticmethod
    def _load_xlsx_data(
        bdata: BytesIO | None = None, filepath: Path | None = None
    ):
        try:
            if filepath:
                wb = openpyxl.load_workbook(filepath, data_only=False)
            elif bdata:
                bdata.seek(0)
                wb = openpyxl.load_workbook(bdata, data_only=False)
            else:
                raise ValueError("No data proveded")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                elif bdata:
                    bdata.seek(0)
                    wb = openpyxl.load_workbook(bdata, data_only=True)
                else:
                    raise ValueError("No data proveded")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheet_tables = []
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_cells = {}
            if with_metadata:
                merged_cells = TableLoader._get_merged_cells_from_xlsx_sheet(sheet)

            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
