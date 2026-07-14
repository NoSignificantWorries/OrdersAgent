import itertools
from dataclasses import Field, dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Self, Set, Tuple, Type

import numpy as np
import openpyxl
import xlrd

from . import config as conf
from . import functional as func
from .config import CellType, HeadersLabels

ValueType = Optional[str | int | Tuple[int, int]]


def cell_classify(value: Optional[str]):
    if value is None or value == "":
        return conf.CellType.EMPTY
    if conf.TYPES_CONFIG.regex is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.regex.items():
            for pattern, groups in patterns:
                if func.check_fullmatch(pattern, value):
                    return cell_type
    if conf.TYPES_CONFIG.fuzzy is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.fuzzy.items():
            for pattern in patterns:
                if func.fuzzy_match(value.lower(), pattern, 70):
                    return cell_type
    return conf.CellType.TEXT


@dataclass
class Cell:
    row: int
    col: int
    value: ValueType = None
    type: CellType = CellType.EMPTY
    merged: bool = False
    merge_parent: Optional[Tuple[int, int]] = None

    def prepare(self) -> "Cell":
        self.type = cell_classify(None if self.value is None else str(self.value))
        return self

    def to_dict(self) -> Dict:
        return {
            "row": self.row,
            "col": self.col,
            "value": self.value,
            "type": self.type.value,
            "merged": self.merged,
            "merge_parent": self.merge_parent,
        }

    def from_dict(self, data: Dict) -> "Cell":
        self.row = data["row"]
        self.col = data["col"]
        self.value = data["value"]
        self.merged = data["merged"]
        self.merge_parent = data["merge_parent"]
        return self


class UniqueIDStorage:
    def __init__(self) -> None:
        self._registry_by_value: Dict[str, int] = {}
        self._registry_by_id: Dict[int, str] = {}
        self._counter = itertools.count(1)

    def add(self, value: str) -> int:
        id = self._registry_by_value.get(value)
        if id is None:
            id = next(self._counter)
            self._registry_by_value[value] = id
            self._registry_by_id[id] = value
        return id

    def get_by_id(self, id: int) -> Optional[str]:
        return self._registry_by_id.get(id)

    def get_by_value(self, value: str) -> Optional[int]:
        return self._registry_by_value.get(value)

    def get_all_by_value(self) -> Dict[str, int]:
        return self._registry_by_value

    def get_all_by_id(self) -> Dict[int, str]:
        return self._registry_by_id

    def get_values_only(self) -> List[str]:
        return list(self._registry_by_value.keys())


class TableV3:
    def __init__(self, nrows: int, ncols: int) -> None:
        self.nrows = nrows
        self.ncols = ncols
        self.cells: List[List[Cell]] = [
            [Cell(row=i, col=j) for j in range(self.ncols)] for i in range(self.nrows)
        ]

        self._table_model = [["."] * self.ncols for _ in range(self.nrows)]

        self._empty_rows = set(range(self.nrows))
        self._empty_cols = set(range(self.ncols))

    def set(
        self,
        row: int,
        col: int,
        value: Optional[str] = None,
        merged: bool = False,
        merge_parent: Optional[Tuple[int, int]] = None,
    ) -> None:
        self.cells[row][col] = Cell(
            row=row, col=col, value=value, merged=merged, merge_parent=merge_parent
        ).prepare()
        if merged:
            self._table_model[row][col] = "*"
        else:
            self._table_model[row][col] = conf.CellTypeLabel[self.cells[row][col].type]
        if self.cells[row][col].type not in CellType.EMPTY and not merged:
            self._empty_cols.discard(col)
            self._empty_rows.discard(row)

    def get_model(self) -> List[str]:
        return ["".join(line) for line in self._table_model]

    def clean_table(self) -> None:
        if bool(self._empty_rows):
            for irow in list(self._empty_rows)[::-1]:
                del self.cells[irow]
                del self._table_model[irow]
            self.nrows -= len(self._empty_rows)

        if bool(self._empty_cols):
            for irow in range(self.nrows):
                for icol in list(self._empty_cols)[::-1]:
                    del self.cells[irow][icol]
                    del self._table_model[irow][icol]
            self.ncols -= len(self._empty_cols)

        self._empty_rows = set()
        self._empty_cols = set()


@dataclass
class ParseResults:
    unique_materials: List[str] = field(default_factory=list)
    materials: List[Optional[int]] = field(default_factory=list)
    x: List[Optional[int]] = field(default_factory=list)
    y: List[Optional[int]] = field(default_factory=list)
    amount: List[Optional[int]] = field(default_factory=list)


class PatternsGroup:
    def __init__(self) -> None:
        self.patterns: Dict[str, Dict[str, int]]
        self.current_window: Optional[List[List[Dict]]] = None
        self.current_data_header: Optional[Dict[int, str]] = None
        self.results: ParseResults = ParseResults()

    def _get_label_counts(
        self, text: str, labels: List[str] | str = conf.HeadersLabels
    ) -> Dict[str, int]:
        counts = {label: 0 for label in labels}
        for label in labels:
            counts[label] += text.count(label)
        return counts

    def _pull_header_indexes(self, row: str, pattern: Dict[str, int]) -> Dict[int, str]:
        columns = {}
        for i, symbol in enumerate(row):
            if symbol in pattern:
                columns[i] = symbol
        return columns

    def _pull_data_by_header(self, header_pattern: Dict[int, str]) -> Dict[str, Dict]:
        results = {}
        for i, label in header_pattern.items():
            results[label] = self.current_window[1][i]
        return results

    def _check_counts(self, pattern: Dict[str, int], data: Dict[str, int]) -> bool:
        for label, count in pattern.items():
            data_count = data.get(label, 0)
            if data_count != count:
                return False
        return True

    def _check_patterns(
        self, counts: Dict[str, int]
    ) -> Optional[Tuple[str, Dict[str, int]]]:
        matched_pattern = None
        for name, pattern in self.patterns.items():
            check_result = self._check_counts(pattern, counts)
            if check_result:
                matched_pattern = (name, pattern)
                break

        return matched_pattern

    def _check_headers(
        self, text: str, headers: List[str] | str = conf.HeadersLabels
    ) -> bool:
        return bool(set(text) & set(headers))

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        raise NotImplementedError("Not implemented method for data window parsing")


class BasicPatterns(PatternsGroup):
    def __init__(self) -> None:
        super().__init__()
        self.patterns = {
            "simple1": {"A": 1, "R": 1, "S": 0, "W": 1, "H": 1, "L": 0},
            "simple2": {"A": 1, "R": 1, "S": 0, "W": 1, "H": 0, "L": 1},
            "simple3": {"A": 1, "R": 1, "S": 0, "W": 0, "H": 1, "L": 1},
            "sized": {"A": 1, "R": 1, "S": 2, "W": 0, "H": 0, "L": 0},
            "sized_uno_duo": {"A": 1, "R": 1, "S": 1, "W": 0, "H": 0, "L": 0},
        }

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        text_window, data_window = window
        self.current_window = data_window
        if self._check_headers(text_window[1]):
            center_counts = self._get_label_counts(text_window[1])
            matched_pattern = self._check_patterns(center_counts)
            if matched_pattern is None:
                if self._check_headers(text_window[0]):
                    top_counts = self._get_label_counts(text_window[0])
                    matched_top = self._check_patterns(top_counts)
                    print(matched_top)
                return

            self.current_data_header = self._pull_header_indexes(
                text_window[1], matched_pattern[1]
            )
            print(matched_pattern[0], text_window, center_counts)
        else:
            if self.current_data_header is not None:
                data = self._pull_data_by_header(self.current_data_header)
                print(data)


class OKNAPatterns(PatternsGroup):
    def __init__(self) -> None:
        super().__init__()
        self.patterns = {
            "data": {"A": 1, "R": 0, "S": 0, "W": 0, "H": 1, "L": 1},
            "material": {"A": 0, "R": 1, "S": 0, "W": 0, "H": 0, "L": 0},
        }
        self._current_material = None

    def _parse_material_row(self, row: str):
        if "t" in row:
            idx = row.index("t")
            material = self.current_window[1][idx]
            self._current_material = material

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        text_window, data_window = window
        self.current_window = data_window
        if self._check_headers(text_window[1]):
            center_counts = self._get_label_counts(text_window[1])
            matched_pattern = self._check_patterns(center_counts)

            if matched_pattern is None:
                return

            if matched_pattern[0] == "material":
                self._parse_material_row(text_window[1])

            self.current_data_header = self._pull_header_indexes(
                text_window[1], matched_pattern[1]
            )
            print(matched_pattern[0], text_window, center_counts)
        else:
            if self.current_data_header is not None:
                data = self._pull_data_by_header(self.current_data_header)
                print(self._current_material, data)


def parser(rows: List[str], data: List[List[Dict]]):
    patterns_list = [BasicPatterns(), OKNAPatterns()]

    for i in range(len(rows)):
        if i == 0:
            window = ["." * len(rows[i])] + rows[i : i + 2]
            data_window = [[]] + data[i : i + 2]
        elif i == len(rows) - 1:
            window = rows[i - 1 :] + ["." * len(rows[i])]
            data_window = data[i - 1 :] + [[]]
        else:
            window = rows[i - 1 : i + 2]
            data_window = data[i - 1 : i + 2]

        for pattern_strategy in patterns_list:
            pattern_strategy.parse_window((window, data_window))


class TableLoader:
    @staticmethod
    def load(bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None):
        if filepath:
            format = filepath.suffix.lower()
            if format == ".xls":
                return TableLoader._load_xls_data(filepath=filepath)
            elif format == ".xlsx":
                return TableLoader._load_xlsx_data(filepath=filepath)
            else:
                raise ValueError("Unsupported file type")
        elif bdata:
            bdata.seek(0)
            header = bdata.read(8)
            bdata.seek(0)

            if header.startswith(b"PK\x03\x04"):
                return TableLoader._load_xlsx_data(bdata=bdata)
            else:
                return TableLoader._load_xls_data(bdata=bdata)
        raise ValueError("No data provided")

    @staticmethod
    def _load_xls_data(
        bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None
    ):
        try:
            if filepath:
                wb = xlrd.open_workbook(str(filepath), formatting_info=True)
            elif bdata:
                bdata.seek(0)
                wb = xlrd.open_workbook(
                    file_contents=bdata.read(), formatting_info=True
                )
            else:
                raise ValueError("Unsupported file type")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = xlrd.open_workbook(str(filepath), formatting_info=False)
                elif bdata:
                    bdata.seek(0)
                    wb = xlrd.open_workbook(
                        file_contents=bdata.read(), formatting_info=False
                    )
                else:
                    raise ValueError("Unsupported file type")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheet_tables = []
        sheets = wb.sheet_names()
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_ranges = []
            if hasattr(sheet, "merged_cells"):
                for merged_cell in sheet.merged_cells:
                    merged_ranges.append(
                        {
                            "min_row": merged_cell[0],
                            "max_row": merged_cell[1] - 1,
                            "min_col": merged_cell[2],
                            "max_col": merged_cell[3] - 1,
                        }
                    )

            st = TableV3(nrows=sheet.nrows, ncols=sheet.ncols)
            cells = []
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)
                    clean_value = func.to_text(value)
                    metadata = {
                        "value": clean_value,
                        "row": row,
                        "col": col,
                        "class": cell_classify(clean_value),
                        "is_merged": False,
                        "merged_info": None,
                    }
                    for mr in merged_ranges:
                        if (mr["min_row"] <= metadata["row"] <= mr["max_row"]) and (
                            mr["min_col"] <= metadata["col"] <= mr["max_col"]
                        ):
                            metadata["is_merged"] = True
                            metadata["merged_info"] = mr
                            break

                    st.set(
                        row,
                        col,
                        clean_value,
                        merged=metadata["is_merged"],
                        merge_parent=metadata["merged_info"],
                    )
                    cells.append(metadata)
            sheet_tables.append(st)
            print(st._empty_cols)
            print(st._empty_rows)
            st.clean_table()
            print("\n".join(st.get_model()))
            print("\n")
        return sheet_tables

    @staticmethod
    def _load_xlsx_data(
        bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None
    ):
        try:
            if filepath:
                wb = openpyxl.load_workbook(filepath, data_only=False)
            elif bdata:
                bdata.seek(0)
                wb = openpyxl.load_workbook(bdata, data_only=False)
            else:
                raise ValueError("No data proveded")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                elif bdata:
                    bdata.seek(0)
                    wb = openpyxl.load_workbook(bdata, data_only=True)
                else:
                    raise ValueError("No data proveded")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheet_tables = []
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_ranges = []
            if with_metadata:
                for merged_range in sheet.merged_cells.ranges:
                    merged_ranges.append(
                        {
                            "min_col": merged_range.min_col - 1,
                            "min_row": merged_range.min_row - 1,
                            "max_col": merged_range.max_col - 1,
                            "max_row": merged_range.max_row - 1,
                        }
                    )

            st = TableV3(nrows=sheet.max_row, ncols=sheet.max_column)
            cells = []
            for row in sheet.iter_rows():
                for cell in row:
                    clean_value = func.to_text(cell.value)

                    metadata = {
                        "value": clean_value,
                        "row": cell.row - 1,
                        "col": cell.column - 1,
                        "class": cell_classify(clean_value),
                        "is_merged": False,
                        "merged_info": None,
                    }

                    for mr in merged_ranges:
                        if (
                            mr["min_row"] <= metadata["row"] <= mr["max_row"]
                            and mr["min_col"] <= metadata["col"] <= mr["max_col"]
                        ):
                            metadata["is_merged"] = True
                            metadata["merged_info"] = mr
                            break

                    st.set(
                        cell.row,
                        cell.col,
                        clean_value,
                        merged=metadata["is_merged"],
                        merge_parent=metadata["merged_info"],
                    )
                    cells.append(metadata)

            sheet_tables.append(st)
            print(st._empty_cols)
            print(st._empty_rows)
            st.clean_table()
            print("\n".join(st.get_model()))
            print("\n")

        return sheet_tables
