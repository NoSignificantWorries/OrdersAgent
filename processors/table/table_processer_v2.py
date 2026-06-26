from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Type

import openpyxl
import xlrd

from . import config as conf
from . import functional as func
from .config import CellType, HeadersLabels


def cell_classify(value: Optional[str]):
    if value is None or value == "":
        return conf.CellType.EMPTY
    if conf.TYPES_CONFIG.regex is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.regex.items():
            for pattern, groups in patterns:
                if func.check_fullmatch(pattern, value):
                    return cell_type
    if conf.TYPES_CONFIG.fuzzy is not None:
        for cell_type, patterns in conf.TYPES_CONFIG.fuzzy.items():
            for pattern in patterns:
                if func.fuzzy_match(value.lower(), pattern, 70):
                    return cell_type
    return conf.CellType.TEXT


class Table:
    def __init__(self) -> None:
        self.cells: List[List[Dict[str, Any]]] = []

        self._current_row = 0
        self._text_row = ""
        self.text_table = []

    @property
    def ncols(self):
        return len(self.cells[0]) if len(self.cells) > 0 else 0

    @property
    def nrows(self):
        return len(self.cells)

    def __repr__(self) -> str:
        return "\n".join(self.text_table)

    def __str__(self) -> str:
        return self.__repr__()

    def add(self, row, col, value: Optional[str], is_merged=False):
        if row > self._current_row:
            self._current_row = row
            self.text_table.append(self._text_row)
            self._text_row = ""
            self.cells.append([])

        if len(self.cells) == 0:
            self.cells.append([])

        cell_type = cell_classify(value)
        self.cells[-1].append(
            {
                "type": cell_type,
                "value": value,
                "row": row,
                "col": col,
                "is_merged": is_merged,
            }
        )
        if value == "" or value is None:
            if is_merged:
                self._text_row += "*"
            else:
                self._text_row += "."
        else:
            self._text_row += conf.CellTypeLabel[cell_type]

    def finish(self):
        self.text_table.append(self._text_row)
        self._text_row = ""

    def get(self, row, col):
        try:
            return self.cells[row][col]
        except Exception:
            return None

    def sparce(self):
        empty_columns = [True for _ in range(self.ncols)]
        empty_rows = [True for _ in range(self.nrows)]
        for irow, row in enumerate(self.text_table):
            for icol, cell in enumerate(row):
                if cell not in ".*":
                    empty_columns[icol] = False
                    empty_rows[irow] = False
        return [i for i, flag in enumerate(empty_rows) if flag], [
            i for i, flag in enumerate(empty_columns) if flag
        ]

    def delete_rows(self, rows: List[int]):
        for irow in rows[::-1]:
            del self.cells[irow]
            del self.text_table[irow]

    def delete_cols(self, cols: List[int]):
        for irow in range(self.nrows):
            for icol in cols[::-1]:
                del self.cells[irow][icol]
                self.text_table[irow] = (
                    self.text_table[irow][:icol] + self.text_table[irow][icol + 1 :]
                )

    def get_text(self):
        return self.text_table


@dataclass
class ParseResults:
    unique_materials: List[str] = field(default_factory=list)
    materials: List[Optional[int]] = field(default_factory=list)
    x: List[Optional[int]] = field(default_factory=list)
    y: List[Optional[int]] = field(default_factory=list)
    amount: List[Optional[int]] = field(default_factory=list)


class PatternsGroup:
    def __init__(self) -> None:
        self.patterns: Dict[str, Dict[str, int]]
        self.current_window: Optional[List[List[Dict]]] = None
        self.current_data_header: Optional[Dict[int, str]] = None
        self.results: ParseResults = ParseResults()

    def _get_label_counts(
        self, text: str, labels: List[str] | str = conf.HeadersLabels
    ) -> Dict[str, int]:
        counts = {label: 0 for label in labels}
        for label in labels:
            counts[label] += text.count(label)
        return counts

    def _pull_header_indexes(self, row: str, pattern: Dict[str, int]) -> Dict[int, str]:
        columns = {}
        for i, symbol in enumerate(row):
            if symbol in pattern:
                columns[i] = symbol
        return columns

    def _pull_data_by_header(self, header_pattern: Dict[int, str]) -> Dict[str, Dict]:
        results = {}
        for i, label in header_pattern.items():
            results[label] = self.current_window[1][i]
        return results

    def _check_counts(self, pattern: Dict[str, int], data: Dict[str, int]) -> bool:
        for label, count in pattern.items():
            data_count = data.get(label, 0)
            if data_count != count:
                return False
        return True

    def _check_patterns(
        self, counts: Dict[str, int]
    ) -> Optional[Tuple[str, Dict[str, int]]]:
        matched_pattern = None
        for name, pattern in self.patterns.items():
            check_result = self._check_counts(pattern, counts)
            if check_result:
                matched_pattern = (name, pattern)
                break

        return matched_pattern

    def _check_headers(
        self, text: str, headers: List[str] | str = conf.HeadersLabels
    ) -> bool:
        return bool(set(text) & set(headers))

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        raise NotImplementedError("Not implemented method for data window parsing")


class BasicPatterns(PatternsGroup):
    def __init__(self) -> None:
        super().__init__()
        self.patterns = {
            "simple1": {"A": 1, "R": 1, "S": 0, "W": 1, "H": 1, "L": 0},
            "simple2": {"A": 1, "R": 1, "S": 0, "W": 1, "H": 0, "L": 1},
            "simple3": {"A": 1, "R": 1, "S": 0, "W": 0, "H": 1, "L": 1},
            "sized": {"A": 1, "R": 1, "S": 2, "W": 0, "H": 0, "L": 0},
            "sized_uno_duo": {"A": 1, "R": 1, "S": 1, "W": 0, "H": 0, "L": 0},
        }

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        text_window, data_window = window
        self.current_window = data_window
        if self._check_headers(text_window[1]):
            center_counts = self._get_label_counts(text_window[1])
            matched_pattern = self._check_patterns(center_counts)
            if matched_pattern is None:
                if self._check_headers(text_window[0]):
                    top_counts = self._get_label_counts(text_window[0])
                    matched_top = self._check_patterns(top_counts)
                    print(matched_top)
                return

            self.current_data_header = self._pull_header_indexes(
                text_window[1], matched_pattern[1]
            )
            print(matched_pattern[0], text_window, center_counts)
        else:
            if self.current_data_header is not None:
                data = self._pull_data_by_header(self.current_data_header)
                print(data)


class OKNAPatterns(PatternsGroup):
    def __init__(self) -> None:
        super().__init__()
        self.patterns = {
            "data": {"A": 1, "R": 0, "S": 0, "W": 0, "H": 1, "L": 1},
            "material": {"A": 0, "R": 1, "S": 0, "W": 0, "H": 0, "L": 0},
        }
        self._current_material = None

    def _parse_material_row(self, row: str):
        if "t" in row:
            idx = row.index("t")
            material = self.current_window[1][idx]
            self._current_material = material

    def parse_window(self, window: Tuple[List[str], List[List[Dict]]]):
        text_window, data_window = window
        self.current_window = data_window
        if self._check_headers(text_window[1]):
            center_counts = self._get_label_counts(text_window[1])
            matched_pattern = self._check_patterns(center_counts)

            if matched_pattern is None:
                return

            if matched_pattern[0] == "material":
                self._parse_material_row(text_window[1])

            self.current_data_header = self._pull_header_indexes(
                text_window[1], matched_pattern[1]
            )
            print(matched_pattern[0], text_window, center_counts)
        else:
            if self.current_data_header is not None:
                data = self._pull_data_by_header(self.current_data_header)
                print(self._current_material, data)


def parser(rows: List[str], data: List[List[Dict]]):
    patterns_list = [BasicPatterns(), OKNAPatterns()]

    for i in range(len(rows)):
        if i == 0:
            window = ["." * len(rows[i])] + rows[i : i + 2]
            data_window = [[]] + data[i : i + 2]
        elif i == len(rows) - 1:
            window = rows[i - 1 :] + ["." * len(rows[i])]
            data_window = data[i - 1 :] + [[]]
        else:
            window = rows[i - 1 : i + 2]
            data_window = data[i - 1 : i + 2]

        for pattern_strategy in patterns_list:
            pattern_strategy.parse_window((window, data_window))


class TableLoader:
    @staticmethod
    def load(bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None):
        if filepath:
            format = filepath.suffix.lower()
            if format == ".xls":
                return TableLoader._load_xls_data(filepath=filepath)
            elif format == ".xlsx":
                return TableLoader._load_xlsx_data(filepath=filepath)
            else:
                raise ValueError("Unsupported file type")
        elif bdata:
            bdata.seek(0)
            header = bdata.read(8)
            bdata.seek(0)

            if header.startswith(b"PK\x03\x04"):
                return TableLoader._load_xlsx_data(bdata=bdata)
            else:
                return TableLoader._load_xls_data(bdata=bdata)
        raise ValueError("No data provided")

    @staticmethod
    def _load_xls_data(
        bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None
    ):
        try:
            if filepath:
                wb = xlrd.open_workbook(str(filepath), formatting_info=True)
            elif bdata:
                bdata.seek(0)
                wb = xlrd.open_workbook(
                    file_contents=bdata.read(), formatting_info=True
                )
            else:
                raise ValueError("Unsupported file type")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = xlrd.open_workbook(str(filepath), formatting_info=False)
                elif bdata:
                    bdata.seek(0)
                    wb = xlrd.open_workbook(
                        file_contents=bdata.read(), formatting_info=False
                    )
                else:
                    raise ValueError("Unsupported file type")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheet_tables = []
        sheets = wb.sheet_names()
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_ranges = []
            if hasattr(sheet, "merged_cells"):
                for merged_cell in sheet.merged_cells:
                    merged_ranges.append(
                        {
                            "min_row": merged_cell[0],
                            "max_row": merged_cell[1] - 1,
                            "min_col": merged_cell[2],
                            "max_col": merged_cell[3] - 1,
                        }
                    )

            st = Table()
            cells = []
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)
                    clean_value = func.to_text(value)
                    metadata = {
                        "value": clean_value,
                        "row": row,
                        "col": col,
                        "class": cell_classify(clean_value),
                        "is_merged": False,
                        "merged_info": None,
                    }
                    for mr in merged_ranges:
                        if (mr["min_row"] <= metadata["row"] <= mr["max_row"]) and (
                            mr["min_col"] <= metadata["col"] <= mr["max_col"]
                        ):
                            metadata["is_merged"] = True
                            metadata["merged_info"] = mr
                            break

                    st.add(row, col, clean_value)
                    cells.append(metadata)
            st.finish()
            deleted_rows, deleted_cols = st.sparce()
            st.delete_rows(deleted_rows)
            st.delete_cols(deleted_cols)
            sheet_tables.append(st)
        return sheet_tables

    @staticmethod
    def _load_xlsx_data(
        bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None
    ):
        try:
            if filepath:
                wb = openpyxl.load_workbook(filepath, data_only=False)
            elif bdata:
                bdata.seek(0)
                wb = openpyxl.load_workbook(bdata, data_only=False)
            else:
                raise ValueError("No data proveded")
            with_metadata = True
        except Exception as first_step_error:
            try:
                if filepath:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                elif bdata:
                    bdata.seek(0)
                    wb = openpyxl.load_workbook(bdata, data_only=True)
                else:
                    raise ValueError("No data proveded")
                with_metadata = False
            except Exception as err:
                print(f"File opening errors: {first_step_error} {err}")
                return None

        sheet_tables = []
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]

            merged_ranges = []
            if with_metadata:
                for merged_range in sheet.merged_cells.ranges:
                    merged_ranges.append(
                        {
                            "min_col": merged_range.min_col - 1,
                            "min_row": merged_range.min_row - 1,
                            "max_col": merged_range.max_col - 1,
                            "max_row": merged_range.max_row - 1,
                        }
                    )

            st = Table()
            cells = []
            for row in sheet.iter_rows():
                for cell in row:
                    clean_value = func.to_text(cell.value)

                    metadata = {
                        "value": clean_value,
                        "row": cell.row - 1,
                        "col": cell.column - 1,
                        "class": cell_classify(clean_value),
                        "is_merged": False,
                        "merged_info": None,
                    }

                    for mr in merged_ranges:
                        if (
                            mr["min_row"] <= metadata["row"] <= mr["max_row"]
                            and mr["min_col"] <= metadata["col"] <= mr["max_col"]
                        ):
                            metadata["is_merged"] = True
                            metadata["merged_info"] = mr
                            break

                    st.add(
                        cell.row - 1,
                        cell.column - 1,
                        clean_value,
                        is_merged=metadata["is_merged"],
                    )
                    cells.append(metadata)

            st.finish()
            deleted_rows, deleted_cols = st.sparce()
            st.delete_rows(deleted_rows)
            st.delete_cols(deleted_cols)
            sheet_tables.append(st)

        return sheet_tables
