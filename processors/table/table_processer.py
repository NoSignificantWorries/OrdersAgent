import re
from collections import defaultdict
from dataclasses import dataclass
from enum import Flag, auto
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Set, Tuple

import openpyxl
import xlrd
from rapidfuzz import fuzz

from materials import ParseResults

from .config import HEADERS_CALLCULATION, MERGES_CALLCULATION


class CellType(Flag):
    TEXT = auto()
    NUMBER = auto()
    SIZE_H = auto()
    LENGTH_H = auto()
    WIDTH_H = auto()
    HEIGHT_H = auto()
    AMOUNT_H = auto()
    MAT_H = auto()
    BARCODE_H = auto()
    MARKING_H = auto()


HEADER = (
    CellType.MAT_H
    | CellType.AMOUNT_H
    | CellType.WIDTH_H
    | CellType.LENGTH_H
    | CellType.HEIGHT_H
    | CellType.SIZE_H
    | CellType.BARCODE_H
    | CellType.MARKING_H
)


@dataclass
class CellTypes:
    regex: Optional[Dict[CellType, List[str]]] = None
    fuzzy: Optional[Dict[CellType, List[str]]] = None


TYPES_CONFIG = CellTypes(
    regex={
        CellType.NUMBER: ["^[+-]?\\d+(?:[.,]\\d+)?$"],
    },
    fuzzy={
        CellType.SIZE_H: [
            "размер",
            "размеры",
            "размеры,мм",
            "размеры[мм]",
            "размеры(мм)",
        ],
        CellType.LENGTH_H: ["длина", "длина,мм", "длина[мм]", "длина(мм)"],
        CellType.WIDTH_H: ["ширина", "ширина,мм", "ширина[мм]", "ширина(мм)"],
        CellType.HEIGHT_H: ["высота", "высота,мм", "высота[мм]", "высота(мм)"],
        CellType.AMOUNT_H: ["кол-во", "количество", "кол-во(шт)", "количество(шт)"],
        CellType.MAT_H: [
            "наименование",
            "обозначение",
            "имя",
            "номенклатура",
            "артикул",
            "типпакета",
            "формула",
            "формулазаполнения",
            "формуласп",
        ],
        CellType.BARCODE_H: ["штрихкод", "шк"],
        CellType.MARKING_H: ["маркировка"],
    },
)


def fuzzy_match(text: str, pattern: str, threshold: int = 50) -> bool:
    val = fuzz.ratio(text, pattern)
    return val >= threshold


class Cell:
    def __new__(
        cls,
        row: int,
        col: int,
        value: str,
        autoclean: bool = False,
        dropempty: bool = False,
        autoclassify: bool = False,
    ):
        if (dropempty and len(str(value).strip()) == 0) or value is None:
            return None
        instance = super().__new__(cls)
        return instance

    def __init__(
        self,
        row: int,
        col: int,
        value: str,
        autoclean: bool = False,
        dropempty: bool = False,
        autoclassify: bool = False,
    ) -> None:
        self._value = str(value)
        self._row = row
        self._col = col
        self._type = CellType.TEXT

        if autoclean:
            self._clear_value()

        if autoclassify:
            self.classify(TYPES_CONFIG)

    def __repr__(self) -> str:
        return f"Cell('{self._value}' {self._row}:{self._col} {self._type})"

    def __str__(self) -> str:
        return self.__repr__()

    @property
    def value(self) -> str:
        return self._value

    @property
    def type(self) -> CellType:
        return self._type

    @property
    def row(self) -> int:
        return self._row

    @property
    def col(self) -> int:
        return self._col

    @property
    def position(self) -> Tuple[int, int]:
        return self._row, self._col

    def set_value(self, value: str) -> None:
        self._value = value

    def _clear_value(self) -> None:
        self._value = self._value.strip()
        self._value = self._value.replace("\xa0", "")

    def classify(self, config: CellTypes) -> CellType:
        value = re.sub(r"\s*", "", self._value)
        if config.regex is not None:
            for cell_type, patterns in config.regex.items():
                for pattern in patterns:
                    if re.fullmatch(pattern, value):
                        self._type = cell_type
                        return cell_type
        if config.fuzzy is not None:
            for cell_type, patterns in config.fuzzy.items():
                for pattern in patterns:
                    if fuzzy_match(value.lower(), pattern, 90):
                        self._type = cell_type
                        return cell_type
        return self._type


class Table:
    @classmethod
    def from_list(cls, cells: List[Cell]) -> "Table":
        new_table = Table()
        for cell in cells:
            new_table.add_cell(cell)
        return new_table

    def __init__(self) -> None:
        self._cells: Dict[Tuple[int, int], Cell] = {}
        self._row_index: Dict[int, Dict[int, Cell]] = defaultdict(dict)
        self._col_index: Dict[int, Dict[int, Cell]] = defaultdict(dict)

        self._cell_index: Dict[CellType, List[Tuple[int, int]]] = {}

        self._rows: Set[int] = set()
        self._cols: Set[int] = set()

    def __repr__(self) -> str:
        return f"Table(len={self.size})"

    def __str__(self) -> str:
        return self.__repr__()

    @property
    def rows(self) -> Iterator:
        return iter(self._rows)

    @property
    def cols(self) -> Iterator:
        return iter(self._cols)

    @property
    def size(self) -> int:
        return len(self._cells)

    @property
    def empty(self) -> bool:
        return self.size == 0

    def get(self, row: int, col: int) -> Optional[Cell]:
        return self._cells.get((row, col))

    def get_row(self, idx: int) -> Optional[Dict[int, Cell]]:
        if idx not in self._rows:
            return None
        return self._row_index[idx]

    def get_col(self, idx: int) -> Optional[Dict[int, Cell]]:
        if idx not in self._cols:
            return None
        return self._col_index[idx]

    def get_by_type(self, type: CellType) -> "Table":
        cells = self._cell_index.get(type)
        if cells is None:
            return Table()

        res = Table()
        for cell_pos in cells:
            res.add_cell(self.get(*cell_pos))

        return res

    def types(self) -> List[CellType]:
        return list(self._cell_index.keys())

    def type_count(self, type: CellType) -> int:
        cells = self._cell_index.get(type)
        if cells is None:
            return 0
        return len(cells)

    def iter_rows(self) -> Iterator[Tuple[int, Dict[int, Cell]]]:
        for row in self._row_index.items():
            yield row

    def iter_cols(self) -> Iterator[Tuple[int, Dict[int, Cell]]]:
        for col in self._col_index.items():
            yield col

    def add_cell(self, cell: Cell) -> None:
        pos = cell.position
        if pos in self._cells:
            self._remove_from_indexes(cell)

        self._cells[cell.position] = cell

        self._row_index[cell.row][cell.col] = cell
        self._col_index[cell.col][cell.row] = cell

        if cell.type not in self._cell_index:
            self._cell_index[cell.type] = []
        self._cell_index[cell.type].append(pos)

        self._rows.add(cell.row)
        self._cols.add(cell.col)

    def _remove_from_indexes(self, cell: Cell) -> None:
        self._row_index[cell.row].pop(cell.col, None)
        self._col_index[cell.col].pop(cell.row, None)

        idx = -1
        for i, subcell_pos in enumerate(self._cell_index[cell.type]):
            if subcell_pos == cell.position:
                idx = i
                break
            if idx >= 0:
                del self._cell_index[cell.type][idx]

    def merge(self, other: "Table") -> None:
        for _, cell in other._cells.items():
            self.add_cell(cell)


header1_1 = [
    CellType.MAT_H,
    CellType.WIDTH_H,
    CellType.HEIGHT_H,
    CellType.AMOUNT_H,
]
header1_2 = [
    CellType.MAT_H,
    CellType.WIDTH_H,
    CellType.LENGTH_H,
    CellType.AMOUNT_H,
]
header1_3 = [
    CellType.MAT_H,
    CellType.LENGTH_H,
    CellType.HEIGHT_H,
    CellType.AMOUNT_H,
]
header_variants = [header1_1, header1_2, header1_3]
optional_headers = [CellType.BARCODE_H, CellType.MARKING_H]
new_header_variants = []
for head in header_variants:
    h1 = head + optional_headers
    h2 = head + [optional_headers[0]]
    h3 = head + [optional_headers[1]]
    h4 = head + optional_headers[::-1]
    new_header_variants += [head, h1, h2, h3, h4]
# print(new_header_variants)
header_variants = sorted(new_header_variants, key=len, reverse=True)

columnrules1 = {
    CellType.MAT_H: CellType.TEXT,
    CellType.WIDTH_H: CellType.NUMBER,
    CellType.HEIGHT_H: CellType.NUMBER,
    CellType.LENGTH_H: CellType.NUMBER,
    CellType.AMOUNT_H: CellType.NUMBER,
    CellType.BARCODE_H: CellType.TEXT | CellType.NUMBER,
    CellType.MARKING_H: CellType.TEXT | CellType.NUMBER,
}


@dataclass
class TableParseResults:
    material: Optional[List[str]] = None
    width: Optional[List[str]] = None
    length: Optional[List[str]] = None
    height: Optional[List[str]] = None
    amount: Optional[List[str]] = None
    barcode: Optional[List[str]] = None
    marking: Optional[List[str]] = None
    size: int = 0

    @property
    def empty(self):
        return self.size == 0 or self.material is None


class TableWorker:
    def __init__(
        self, bdata: Optional[BytesIO] = None, filepath: Optional[Path] = None
    ) -> None:
        self.bdata = bdata
        self.filepath = filepath
        if self.filepath:
            self.format = self.filepath.suffix
        self.tables = None

    def open_and_clean(self):
        if self.format == ".xls":
            self._open_and_clean_xls()
        elif self.format == ".xlsx":
            self._open_and_clean_xlsx()
        else:
            raise ValueError("Unsupported file type")

    def _open_and_clean_xls(self):
        try:
            if self.bdata:
                wb = xlrd.open_workbook(
                    file_contents=self.bdata.read(), formatting_info=False
                )
            else:
                wb = xlrd.open_workbook(str(self.filepath), formatting_info=False)
        except Exception:
            return ValueError("File damaged or not exists")

        sheets = wb.sheet_names()
        sheet_tables = []
        for sheetname in sheets:
            sheet = wb[sheetname]

            table = Table()
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    val = sheet.cell_value(row, col)
                    cell = Cell(
                        row, col, val, dropempty=True, autoclean=True, autoclassify=True
                    )
                    if cell is None:
                        continue
                    # print(cell)
                    table.add_cell(cell)
            sheet_tables.append(table)
        self.tables = sheet_tables

    def _open_and_clean_xlsx(self):
        try:
            if self.bdata:
                wb = openpyxl.load_workbook(self.bdata)
            else:
                wb = openpyxl.load_workbook(str(self.filepath))
        except Exception:
            return ValueError("File damaged or not exists")

        sheets = wb.sheetnames
        sheet_tables = []
        for sheetname in sheets:
            sheet = wb[sheetname]

            table = Table()
            for row, row_obj in enumerate(sheet.values):
                for col, val in enumerate(row_obj):
                    # val = sheet.cell_value(row, col)
                    cell = Cell(
                        row, col, val, dropempty=True, autoclean=True, autoclassify=True
                    )
                    if cell is None:
                        continue
                    # print(cell)
                    table.add_cell(cell)
            sheet_tables.append(table)
        self.tables = sheet_tables

    def simple_parser(self):
        if self.tables is None or not bool(self.tables):
            raise ValueError("No Table data, open file firstly")

        def _parse_sheet(table: Table):
            if table.empty:
                return TableParseResults()

            headers = Table()
            for htype in columnrules1.keys():
                headers.merge(table.get_by_type(htype))
            # print(headers)
            # print(headers._cells)
            # print(list(headers.rows))
            # print(list(headers.cols))
            if headers.empty:
                return TableParseResults()
            variant_matched = False
            match_variant = None
            target_indexes = {}
            for variant in header_variants:
                this_variant = True
                target_indexes = {}
                for idx, row in headers.iter_rows():
                    types = [cl.type for _, cl in row.items()]
                    if len(row) != len(variant) or types != variant:
                        this_variant = False
                        break
                    indexes = [pos for pos, _ in row.items()]
                    target_indexes[idx] = indexes
                if this_variant:
                    variant_matched = True
                    match_variant = variant
                    break

            if not variant_matched or match_variant is None:
                return TableParseResults()

            # print(variant_matched, match_variant, target_indexes)
            results = {header: [] for header in match_variant}

            h_idx = list(target_indexes.keys())[0]
            cols = target_indexes[h_idx]
            for idx, row in table.iter_rows():
                if idx <= h_idx:
                    continue
                values_row = {header: None for header in match_variant}
                for col, htype in zip(cols, match_variant):
                    cell = table.get(idx, col)
                    values_row[htype] = cell
                correct_row = True
                for header, match in values_row.items():
                    if match is None:
                        correct_row = False
                        break
                    if match.type not in columnrules1[header]:
                        correct_row = False
                        break

                if correct_row:
                    for header, match in values_row.items():
                        results[header].append(match)

            # print(results)
            res = TableParseResults()
            for key, cells in results.items():
                values = list(map(lambda c: c.value, cells))
                match key:
                    case CellType.MAT_H:
                        res.material = values
                    case CellType.WIDTH_H:
                        res.width = values
                    case CellType.LENGTH_H:
                        res.length = values
                    case CellType.HEIGHT_H:
                        res.height = values
                    case CellType.AMOUNT_H:
                        res.amount = values
                    case CellType.BARCODE_H:
                        res.barcode = values
                    case CellType.MARKING_H:
                        res.marking = values
            if res.material is None:
                return TableParseResults()
            res.size = len(res.material)
            return res

        results = []
        for sheet in self.tables:
            results.append(_parse_sheet(sheet))

        return results


def make_request_xlsx(
    origin_table: List[TableParseResults], elements: Dict[str, ParseResults]
):
    wb = openpyxl.Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    all_empty = True
    for i, sheet_data in enumerate(origin_table):
        ws = wb.create_sheet(title=f"Sheet {i}")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
        ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
        for col_idx, header_content in [
            (10, "Характеристики"),
            (13, "Доп. информация"),
        ]:
            cell = ws.cell(row=1, column=col_idx + 1, value=header_content)

        for col_idx, header_content in enumerate(
            [
                "Номенклатура 1С",
                "П1",
                "Р1",
                "П2",
                "Р2",
                "П3",
                "Р3",
                "П4",
                "Герметик",
                "Тип изделия",
                "X",
                "Y",
                "Кол-во",
                "Маркировка",
                "ПЗ",
                "ШК",
                "Номер фигуры",
                "Уточнения",
            ]
        ):
            cell = ws.cell(row=2, column=col_idx + 1, value=header_content)

        ws.row_dimensions[1].height = 15
        ws.row_dimensions[2].height = 25

        if sheet_data.empty or sheet_data.material is None or sheet_data.amount is None:
            print("Empty sheet")
            continue

        X, Y = None, None
        if sheet_data.width is None:
            X = sheet_data.length
            Y = sheet_data.height
        elif sheet_data.length is None:
            X = sheet_data.width
            Y = sheet_data.height
        elif sheet_data.height is None:
            X = sheet_data.width
            Y = sheet_data.length

        if X is None or Y is None:
            print("No enough sides on sheet")
            continue

        all_empty = False

        current_row = 3
        for obj_i in range(sheet_data.size):
            material, x, y, amount = (
                sheet_data.material[obj_i],
                X[obj_i],
                Y[obj_i],
                sheet_data.amount[obj_i],
            )
            barcode = ""
            if sheet_data.barcode is not None:
                barcode = sheet_data.barcode[obj_i]
            marking = ""
            if sheet_data.marking is not None:
                marking = sheet_data.marking[obj_i]
            for i, (target, _, _) in enumerate(
                elements[material].matches,
                start=2,
            ):
                cell = ws.cell(row=current_row, column=i, value=target)

            x = x.replace(",", ".")
            y = y.replace(",", ".")
            amount = amount.replace(",", ".")
            cell = ws.cell(row=current_row, column=1, value=material)
            cell = ws.cell(row=current_row, column=11, value=int(float(x)))
            cell = ws.cell(row=current_row, column=12, value=int(float(y)))
            cell = ws.cell(row=current_row, column=13, value=int(float(amount)))
            cell = ws.cell(row=current_row, column=14, value=marking)
            cell = ws.cell(row=current_row, column=16, value=barcode)
            cell = ws.cell(
                row=current_row,
                column=18,
                value=elements[material].postfix,
            )
            current_row += 1

    if all_empty:
        return None
    return wb


def make_callculation_xlsx(
    origin_table: List[TableParseResults], elements: Dict[str, ParseResults]
):
    wb = openpyxl.Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    all_empty = True
    current_row = 3
    ws = wb.create_sheet(title="Sheet1")
    for i, sheet_data in enumerate(origin_table):
        for (row_idx, col_idx), header_content in HEADERS_CALLCULATION.items():
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=header_content)

        for m_range in MERGES_CALLCULATION:
            minc, minr, maxc, maxr = m_range
            ws.merge_cells(
                start_row=minr,
                start_column=minc,
                end_row=maxr,
                end_column=maxc,
            )

        if sheet_data.empty or sheet_data.material is None or sheet_data.amount is None:
            print("Empty sheet")
            continue

        X, Y = None, None
        if sheet_data.width is None:
            X = sheet_data.length
            Y = sheet_data.height
        elif sheet_data.length is None:
            X = sheet_data.width
            Y = sheet_data.height
        elif sheet_data.height is None:
            X = sheet_data.width
            Y = sheet_data.length

        if X is None or Y is None:
            print("No enough sides on sheet")
            continue

        all_empty = False

        for obj_i in range(sheet_data.size):
            material, x, y, amount = (
                sheet_data.material[obj_i],
                X[obj_i],
                Y[obj_i],
                sheet_data.amount[obj_i],
            )
            for i, (_, article, _) in enumerate(
                elements[material].matches,
                start=3,
            ):
                cell = ws.cell(row=current_row, column=i, value=article)

            x = x.replace(",", ".")
            y = y.replace(",", ".")
            amount = amount.replace(",", ".")
            cell = ws.cell(row=current_row, column=1, value=str(obj_i + 1))
            cell = ws.cell(row=current_row, column=2, value=int(float(amount)))
            cell = ws.cell(row=current_row, column=10, value=int(float(x)))
            cell = ws.cell(row=current_row, column=11, value=int(float(y)))
            cell = ws.cell(row=current_row, column=17, value="YES")
            current_row += 1

    if all_empty:
        return None
    return wb
