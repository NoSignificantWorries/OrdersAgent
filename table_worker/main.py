from io import BytesIO
from pathlib import Path
from typing import Dict, List

import openpyxl

from cloud import MinIOClient, get_bytes_object, put_bytes_object
from database import DatabaseManager, MaterialRepository, init_database
from materials import DELIMETERS, ParseResults, ParserV2
from table import TableParseResults, TableWorker

ATTACHMENTS_BUCKET = "orders-attachments"
RESULTS_BUCKET = "results"


def make_xlsx(origin_table: List[TableParseResults], elements: Dict[str, ParseResults]):
    wb = openpyxl.Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    all_empty = True
    for i, sheet_data in enumerate(origin_table):
        ws = wb.create_sheet(title=f"Sheet {i}")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
        ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
        for col_idx, header_content in [
            (10, "Характеристики"),
            (13, "Доп. информация"),
        ]:
            cell = ws.cell(row=1, column=col_idx + 1, value=header_content)

        for col_idx, header_content in enumerate(
            [
                "Номенклатура 1С",
                "П1",
                "Р1",
                "П2",
                "Р2",
                "П3",
                "Р3",
                "П4",
                "Герметик",
                "Тип изделия",
                "X",
                "Y",
                "Кол-во",
                "Маркировка",
                "ПЗ",
                "ШК",
                "Номер фигуры",
                "Уточнения",
            ]
        ):
            cell = ws.cell(row=2, column=col_idx + 1, value=header_content)

        ws.row_dimensions[1].height = 15
        ws.row_dimensions[2].height = 25

        if sheet_data.empty or sheet_data.material is None or sheet_data.amount is None:
            print("Empty sheet")
            continue

        X, Y = None, None
        if sheet_data.width is None:
            X = sheet_data.length
            Y = sheet_data.height
        elif sheet_data.length is None:
            X = sheet_data.width
            Y = sheet_data.height
        elif sheet_data.height is None:
            X = sheet_data.width
            Y = sheet_data.length

        if X is None or Y is None:
            print("No enough sides on sheet")
            continue

        all_empty = False

        current_row = 3
        for obj_i in range(sheet_data.size):
            material, x, y, amount = (
                sheet_data.material[obj_i],
                X[obj_i],
                Y[obj_i],
                sheet_data.amount[obj_i],
            )
            for i, part in enumerate(
                elements[material].matches,
                start=2,
            ):
                cell = ws.cell(row=current_row, column=i, value=part)

            cell = ws.cell(row=current_row, column=1, value=material)
            cell = ws.cell(row=current_row, column=11, value=x)
            cell = ws.cell(row=current_row, column=12, value=y)
            cell = ws.cell(row=current_row, column=13, value=amount)
            cell = ws.cell(
                row=current_row,
                column=18,
                value=elements[material].postfix,
            )
            current_row += 1

    if all_empty:
        return None
    return wb


def main() -> None:
    try:
        init_database()
        cloud_client = MinIOClient.get_client()

    finally:
        ...


def development() -> None:
    init_database()
    cloud_client = MinIOClient.get_client()

    # testfile = Path("../private/tables/1108A.xls")
    filename = "033/1108A.xls"

    file_data = get_bytes_object(cloud_client, ATTACHMENTS_BUCKET, filename)
    print(file_data)
    # try:
    #     response = cloud_client.get_object(ATTACHMENTS_BUCKET, filename)
    #     file_data = BytesIO(response.read())
    #     response.close()
    #     response.release_conn()
    # except Exception:
    #     print("Errors while file reading")
    #     return

    # worker = TableWorker(None, testfile)
    worker = TableWorker(file_data, Path(filename))
    worker.open_and_clean()
    if worker.tables is None or not bool(worker.tables):
        print("Errors with table")

    res = worker.simple_parser()
    print(res)

    unique_materials = set()
    for table in res:
        if table.empty:
            continue
        unique_materials |= set(table.material)
    unique_materials = list(unique_materials)
    print(unique_materials)

    parser = ParserV2(DELIMETERS)
    unique_materials_dict = {}
    unique_parts = set()
    for material in unique_materials:
        parse_results = parser.parse(material)
        unique_materials_dict[material] = parse_results
        unique_parts |= set(parse_results.parts)
    unique_parts = list(unique_parts)

    print(unique_materials_dict)
    print(unique_parts)

    material_repo = MaterialRepository()
    matches = material_repo.batch_find(unique_parts)
    print(matches)
    questions = []
    for part, mat_match in matches.items():
        if mat_match is None:
            questions.append({part: False})
            continue
        _, bl = mat_match
        if bl:
            questions.append({part: True})
            continue
    print(questions)

    if questions:
        answers = {"4HPBronze20": ("Bronze20", False), "14": ("14Mr", False)}

        answers_flat = [(part, data[0], data[1]) for part, data in answers.items()]

        material_repo.batch_add(answers_flat)

        matches = material_repo.batch_find(unique_parts)

    print(matches)

    for material, material_obj in unique_materials_dict.items():
        for part in material_obj.parts:
            material_obj.matches.append(matches[part][0])
    print(unique_materials_dict)

    wb = make_xlsx(res, unique_materials_dict)
    print(wb)

    if wb is None:
        print("Empty file")
        return

    data = BytesIO()
    wb.save(data)
    data.seek(0)

    err = put_bytes_object(
        cloud_client,
        RESULTS_BUCKET,
        filename,
        data,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    print(err)
    # try:
    #     cloud_client.put_object(
    #         bucket_name=RESULTS_BUCKET,
    #         object_name=filename,
    #         data=data,
    #         length=data.getbuffer().nbytes,
    #         content_type=,
    #     )
    # except Exception:
    #     print("Error while saving file in the cloud")

    DatabaseManager.close()


if __name__ == "__main__":
    # main()
    development()
