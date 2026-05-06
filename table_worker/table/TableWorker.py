import asyncio
import json
import re
from functools import partial
from html import escape
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple, Union

# data working
import numpy as np

# tables processing
import openpyxl
import pandas as pd
import tqdm
import xarray as xr
import xlrd
from rapidfuzz import fuzz, process
from sqlalchemy import column

from .MaterialParser import (
    DELIMETERS,
    DatabaseManager,
    MaterialProcessor,
    ParseResults,
    ParserV2,
    initialize_app,
)

_basic_table = """
<!DOCTYPE html>
<html lang="ru">
  <head>
    <meta charset="UTF-8">
    <title>File: {filename}</title>
    <style>
    table {{ border-collapse: collapse; font-family: sans-serif; font-size: 12px; }}
    td, th {{ border: 1px solid #000; padding: 2px 4px; max-width: 200px; overflow: hidden; white-space: nowrap; }}
    .cell-number    {{ background-color: #e0ffe0; }}
    .cell-text {{ background-color: #e0f0ff; }}
    .cell-barcode {{ background-color: #f5c2e7; }}
    .cell-maybe-material {{ background-color: #74c7ec; }}
    .cell-empty    {{ background-color: #666; color: #222; }}
    .cell-size {{ background-color: #00ff00; }}
    .cell-width {{ background-color: #00ffff; }}
    .cell-height {{ background-color: #0f00f0; }}
    .cell-length {{ background-color: #ff00ff; }}
    .cell-amount {{ background-color: #ffff00; }}
    .cell-material {{ background-color: #00ffff; }}
    </style>
  </head>
  <body>
    <table>{table}</table>
  </body>
</html>
"""


def _get_filename(path: Path):
    filename = "".join(path.name.split(".")[:-1])
    return filename


def _make_table_from_xa(da: xr.DataArray) -> List[str]:
    sheets_html = []

    values = da.sel(attr="value")
    classes = da.sel(attr="class").values

    def process_value(x):
        if pd.isna(x):
            return "NULL"
        return str(x)

    vectorized_value = np.vectorize(process_value, otypes=[object])
    vectorized_escape = np.vectorize(lambda x: escape(str(x)), otypes=[object])

    processed_values = vectorized_value(values.values)
    escaped_values = vectorized_escape(processed_values)

    sh, r, c, _ = da.shape

    for i in range(sh):
        html_rows = []
        for j in range(r):
            tds = []
            for k in range(c):
                td = f"<td class='cell-{classes[i, j, k]}'>{escaped_values[i, j, k]}"
                tds.append(td)
            html_rows.append("<tr>" + "".join(tds) + "</tr>")
        sheets_html.append("".join(html_rows))

    return sheets_html


def _create_tables(da: xr.DataArray, dirpath: Path, filename: str) -> None:
    tables_html = _make_table_from_xa(da)

    for i, table in enumerate(tables_html):
        filepath = dirpath / Path(f"{filename}_sheet_{i}.html")
        content = _basic_table.format(table=table, filename=f"{filename}_sheet_{i}")
        dirpath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w") as file:
            file.write(content)


class ColumnsConfig:
    def __init__(self, config_path):
        self.path = config_path

        with open(self.path, "r") as file:
            data = json.load(file)

        self.col_asc = data["column-associations"]
        self.cell_regex = data["regex-classes"]


class TableExtruder:
    def __init__(self, name: str) -> None:
        self.name = name
        self.header_patterns = None
        self.sheets_data = None

    def _extrude_header_from_row(self, row):
        if self.header_patterns is None:
            raise ValueError("[ERROR]: Empty headers list.")

    def process_data(self, table: xr.DataArray):
        raise NotImplementedError("[ERROR]: Not implemented method 'find pattern'")


class StandartExtruder(TableExtruder):
    def __init__(self) -> None:
        super().__init__(name="Standart")
        self.header_patterns = [
            {"material": 1, "width": 1, "height": 1, "amount": 1},
            {"material": 1, "length": 1, "height": 1, "amount": 1},
            {"material": 1, "width": 1, "length": 1, "amount": 1},
            {"material": 1, "size": 2, "amount": 1},
        ]
        self.header_classes = {
            "material": "text",
            "size": "number",
            "length": "number",
            "width": "number",
            "height": "number",
            "amount": "number",
        }
        self.sheets_data = None

    def _make_xlsx(self, data):
        wb = openpyxl.Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

        all_empty = True
        for i, sheet_data in enumerate(data):
            ws = wb.create_sheet(title=f"Sheet {i}")

            headers = sheet_data["header"]
            sheet_data_data = sheet_data["data"]

            if headers is None:
                print("[WARN]: Not a format on sheet")
                continue
            else:
                all_empty = False

            for col_idx, header_name in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header_name)

            current_row = 2
            for line in sheet_data_data:
                for col_idx, value in enumerate(line, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                current_row += 1

        if all_empty:
            return None
        return wb

    def process_data(self, table: xr.DataArray):
        def check_row(row):
            for i, header_pattern in enumerate(self.header_patterns):
                all_correct = True
                indexes = []
                keys = []
                for key, cnt in header_pattern.items():
                    mask = np.where(row == key)[0]
                    if len(mask) != cnt:
                        all_correct = False
                        break
                    indexes.append(mask[0])
                    keys.append(key)
                if all_correct:
                    return (True, indexes, keys)
            return (False, [], [])

        def check_classes(row, pattern_idx, pattern):
            for i, key in zip(pattern_idx, pattern):
                if row[i] != self.header_classes[key]:
                    return False
            return True

        def process_sheet(sheet):
            is_pattern = False
            pattern = None
            pattern_idx = None
            table = {"header": None, "data": []}
            for i, row in enumerate(sheet):
                vals = row.sel(attr="value")
                clss = row.sel(attr="class")

                pattern_here, _pattern_idx, _pattern = check_row(clss)
                if pattern_here:
                    is_pattern = True
                    if _pattern == pattern:
                        raise ValueError("[ERROR]: Wrong data format")
                    pattern = _pattern
                    pattern_idx = _pattern_idx
                    table["header"] = pattern
                    continue
                if is_pattern:
                    if check_classes(clss, pattern_idx, pattern):
                        table["data"].append(vals[pattern_idx].values)

            return table

        sh_tables = []
        for sheet in table:
            sh_tables.append(process_sheet(sheet))

        self.sheets_data = sh_tables.copy()

        return self._make_xlsx(sh_tables)


def fuzzy_match(text, pattern, threshold=50) -> bool:
    val = fuzz.ratio(text, pattern)
    return val >= threshold


def match_class(cell, config: ColumnsConfig) -> str:
    if pd.isna(cell):
        return "empty"
    cell = str(cell)
    cell = re.sub(r"\s*", "", cell)

    for cls, pattern in config.cell_regex.items():
        if re.fullmatch(pattern, cell):
            return cls

    for cls, associations in config.col_asc.items():
        for asc in associations:
            if fuzzy_match(cell.lower(), asc, 90):
                return cls

    return "text"


def match_classes_to_df(data: pd.DataFrame, config: ColumnsConfig) -> pd.DataFrame:
    classes = pd.DataFrame(index=data.index, columns=data.columns)
    match_class_with_config = partial(match_class, config=config)
    for col_idx in data.columns:
        col = data[col_idx].astype(str)
        classes[col_idx] = col.apply(match_class_with_config)
    return classes


async def ask_user_about_materials(materials):
    pass


class TableWorker:
    def __init__(
        self, data: Optional[BytesIO], filepath: Path, config: ColumnsConfig
    ) -> None:
        self.bytes_data = data
        self.path = filepath
        self.format = filepath.suffix
        self.name = _get_filename(filepath)
        self.attr = ["value", "class"]
        self.data = None
        self.parsed_data = None
        self.parsed_materials = None
        self.wb = None
        self.origin_shapes = []

        self.config = config

    def open_and_clean(self):
        if self.format == ".xls":
            self._open_and_clean_xls()
        elif self.format == ".xlsx":
            self._open_and_clean_xlsx()

    def _open_and_clean_xls(self):
        try:
            if self.bytes_data:
                wb = xlrd.open_workbook(
                    file_contents=self.bytes_data.read(), formatting_info=False
                )
            else:
                wb = xlrd.open_workbook(str(self.path), formatting_info=False)
        except Exception:
            return

        sheets_data = []
        sheets = wb.sheet_names()
        for sheetname in sheets:
            sheet = wb[sheetname]

            data = [
                [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                for row in range(sheet.nrows)
            ]

            df = pd.DataFrame(data)
            df = df.astype(str)
            df = df.astype(str).replace("nan", np.nan)
            df = df.replace(r"^\s*$", np.nan, regex=True)

            df.dropna(axis=0, how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df = df.reset_index(drop=True)

            if not df.empty:
                self.origin_shapes.append(df.shape)
                sheets_data.append(df)

        self._sheets_to_xarray(sheets_data)

    def _open_and_clean_xlsx(self):
        try:
            if self.bytes_data:
                wb = openpyxl.load_workbook(self.bytes_data)
            else:
                wb = openpyxl.load_workbook(self.path)
        except Exception as err:
            return

        sheets_data = []
        sheets = wb.sheetnames
        for sheetname in sheets:
            sheet = wb[sheetname]

            data = sheet.values
            df = pd.DataFrame(data)
            df = df.astype(str)
            df = df.astype(str).replace("nan", np.nan)
            df = df.replace(r"^\s*$", np.nan, regex=True)

            df.dropna(axis=0, how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df = df.reset_index(drop=True)

            if not df.empty:
                self.origin_shapes.append(df.shape)
                sheets_data.append(df)

        self._sheets_to_xarray(sheets_data)

    def _sheets_to_xarray(self, sheets):
        if not bool(self.origin_shapes):
            return

        if not bool(sheets):
            return

        max_rows = max(shape[0] for shape in self.origin_shapes)
        max_cols = max(shape[1] for shape in self.origin_shapes)

        n = len(sheets)
        aligned_data = np.full(
            (n, max_rows, max_cols, len(self.attr)), np.nan, dtype=object
        )

        for i, (sheet_data, (n_rows, n_cols)) in enumerate(
            zip(sheets, self.origin_shapes)
        ):
            aligned_data[i, :n_rows, :n_cols, 0] = sheet_data

        self.data = xr.DataArray(
            aligned_data,
            dims=["sheet", "row", "column", "attr"],
            coords={
                "sheet": [f"sheet_{i}" for i in range(n)],
                "row": range(max_rows),
                "column": range(max_cols),
                "attr": self.attr,
            },
        )

    def match_class_to_data(self):
        if self.data is None:
            return

        values = self.data.sel(attr="value")

        conf_func = partial(match_class, config=self.config)
        vectorized_func = np.vectorize(conf_func, otypes=[object])

        class_labels = vectorized_func(values.values)

        self.data[..., self.attr.index("class")] = class_labels

    def apply_extruder(self, extruder: TableExtruder):
        if self.data is None:
            # raise ValueError("[ERROR]: Empty table data.")
            print("[ERROR]: Empy table data")
            return None

        self.wb = extruder.process_data(self.data)
        if self.wb is None:
            return None

        self.parsed_data = extruder.sheets_data
        return extruder.sheets_data

    async def parse_materials(self, processor: MaterialProcessor):
        if self.parsed_data is None:
            return None

        materials = dict()
        for sheet in self.parsed_data:
            if sheet["header"] is None:
                print("No data on sheet")
                continue
            material_column_idx = sheet["header"].index("material")
            for i, row in enumerate(sheet["data"]):
                material = row[material_column_idx].strip()
                row[material_column_idx] = material
                if material not in materials:
                    materials[material] = []
                materials[material].append(i)

        # print(materials)

        material_questions = dict()
        parsed_materials = dict()
        is_problem = False
        for material in materials.keys():
            parsed = await processor.process_line(material)
            if material not in parsed_materials:
                parsed_materials[material] = {
                    "postfix": parsed.postfix,
                    "parts": dict(),
                }
            for i, (p, m) in enumerate(zip(parsed.parts, parsed.matches)):
                parsed_materials[material]["parts"][p] = m
                if m is None:
                    is_problem = True
                    if material not in material_questions:
                        material_questions[material] = []
                    material_questions[material].append((p, i))
        print(material_questions)
        print(parsed_materials)
        if is_problem:
            await ask_user_about_materials(material_questions)

        self.parsed_materials = parsed_materials

    def _make_xlsx(self):
        if self.parsed_data is None or self.parsed_materials is None:
            return None

        wb = openpyxl.Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

        all_empty = True
        for i, sheet_data in enumerate(self.parsed_data):
            ws = wb.create_sheet(title=f"Sheet {i}")

            headers = sheet_data["header"]
            sheet_data_data = sheet_data["data"]

            if headers is None:
                print("[WARN]: Not a format on sheet")
                continue
            else:
                all_empty = False

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
            ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
            for col_idx, header_content in [
                (10, "Характеристики"),
                (13, "Доп. информация"),
            ]:
                cell = ws.cell(row=1, column=col_idx + 1, value=header_content)

            for col_idx, header_content in enumerate(
                [
                    "Номенклатура 1С",
                    "П1",
                    "Р1",
                    "П2",
                    "Р2",
                    "П3",
                    "Р3",
                    "П4",
                    "Герметик",
                    "Тип изделия",
                    "X",
                    "Y",
                    "Кол-во",
                    "Маркировка",
                    "ПЗ",
                    "ШК",
                    "Номер фигуры",
                    "Уточнения",
                ]
            ):
                cell = ws.cell(row=2, column=col_idx + 1, value=header_content)

            ws.row_dimensions[1].height = 15
            ws.row_dimensions[2].height = 25

            current_row = 3
            for line in sheet_data_data:
                material, x, y, amount = line
                for i, (_, matched) in enumerate(
                    self.parsed_materials[material]["parts"].items(), start=2
                ):
                    if matched:
                        value, black_list = matched
                        cell = ws.cell(row=current_row, column=i, value=value)
                    else:
                        cell = ws.cell(row=current_row, column=i, value=matched)

                cell = ws.cell(row=current_row, column=1, value=material)
                cell = ws.cell(row=current_row, column=11, value=x)
                cell = ws.cell(row=current_row, column=12, value=y)
                cell = ws.cell(row=current_row, column=13, value=amount)
                cell = ws.cell(
                    row=current_row,
                    column=18,
                    value=self.parsed_materials[material]["postfix"],
                )
                current_row += 1

        if all_empty:
            return None
        self.wb = wb
        return wb

    def save_wb(self, path: Path):
        if self.wb is None:
            return
        path = path / Path(f"{self.name}.xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)


async def development_async():
    await initialize_app()

    private_dir = Path("~/Projects/OrdersAgent/private").expanduser()
    tables_path = private_dir / Path("tables")
    result_path = private_dir / Path("results/tables")
    result_html_path = private_dir / Path("results/html")

    config_path = Path("../config.json").resolve()
    conf = ColumnsConfig(config_path)

    pipeline = ParserV2(DELIMETERS)
    processor = MaterialProcessor(pipeline)

    # table_path = tables_path / Path("BTs_Kirova_steklopakety.xlsx")
    # table_path = tables_path / Path("1108A.xls")
    all_materials = []
    count_of_files = 0
    parsed_files = 0
    no_parsed_files = []
    strange_files = []
    for table_path in tables_path.iterdir():
        print("==>", table_path.name)
        count_of_files += 1
        table = TableWorker(None, table_path, conf)

        table.open_and_clean()

        table.match_class_to_data()

        data = table.apply_extruder(StandartExtruder())

        if data is not None:
            for sheet in data:
                if sheet["header"] is None:
                    print("[WARN]: Empty sheet")
                    continue

                await table.parse_materials(processor)

                table._make_xlsx()

                idx = sheet["header"].index("material")
                materials = []
                for line in sheet["data"]:
                    materials.append(line[idx].strip())
                materials = list(set(materials))
                all_materials += materials
                # print(*materials, sep="\n")
            parsed_files += 1
            table.save_wb(result_path)
        else:
            no_parsed_files.append(table_path.name)
            print("ERROR: No data")

        if table.data is None:
            print(f"Empty Table for file ((({table_path})))")
            strange_files.append(table_path.name)
        else:
            _create_tables(table.data, result_html_path, table.name)

    await DatabaseManager.close()

    if count_of_files == 0:
        print("No files")
    else:
        print(parsed_files, count_of_files)
        print("Total percent:", parsed_files / count_of_files * 100)
    with open("res.txt", "w") as file:
        file.write("\n".join(all_materials))
    with open("no_parsed.txt", "w") as file:
        file.write("\n".join(no_parsed_files) + "\n" + "\n".join(strange_files))


def development() -> None:
    asyncio.run(development_async())


if __name__ == "__main__":
    development()
